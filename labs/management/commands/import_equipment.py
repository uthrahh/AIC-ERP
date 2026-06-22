from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime

from labs.models import EquipmentMaster


class Command(BaseCommand):

    def handle(self, *args, **kwargs):

        file_path = "media/lab_imports/equipment_master.xlsx"

        wb = load_workbook(file_path)
        ws = wb.active

        EquipmentMaster.objects.all().delete()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[3]:
                continue
            procurement_date = None
            if row[6]:
                    amount = str(row[6])
                    amount = amount.replace("₹", "")
                    amount = amount.replace(",", "")
                    amount = amount.strip()
            if row[5]:
                try:
                    procurement_date = datetime.strptime(
                        str(row[5]),
                        "%d.%m.%Y"
                    ).date()
                except Exception:
                    pass

            amount = 0

            if isinstance(amount, str):
                amount = (
                    amount.replace("₹", "")
                        .replace(",", "")
                        .strip()
                )

            EquipmentMaster.objects.create(
                applicant_centre=row[1] or "",
                pan_number=row[2] or "",
                equipment_name=row[3] or "",
                brand_name=row[4] or "",
                procurement_date=procurement_date,
                amount=amount,
            )

        self.stdout.write(
            self.style.SUCCESS(
                "Equipment imported successfully"
            )
        )